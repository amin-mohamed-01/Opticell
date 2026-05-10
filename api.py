from fastapi import FastAPI
import os
from openpyxl import load_workbook

app = FastAPI()

DATA_TRACK_DIR = "data_track"


@app.get("/")
def home():
    return {
        "message": "API is running",    
        "endpoints": ["/get_errors"],
        "description": "Use /get_errors to fetch organized error reports"
    }


@app.get("/get_errors")
def get_errors():
    final_output = []

    for filename in os.listdir(DATA_TRACK_DIR):
        if filename.endswith(".xlsx"):
            filepath = os.path.join(DATA_TRACK_DIR, filename)

            wb = load_workbook(filepath)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]

            # Ensure required columns exist
            if "Error_Type" not in headers or "Reason" not in headers:
                continue

            error_idx = headers.index("Error_Type")
            reason_idx = headers.index("Reason")

            details = []
            summary = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                error_type = row[error_idx]
                reason = row[reason_idx]

                # Add details
                details.append({
                    "error_type": error_type,
                    "reason": reason
                })

                # Count error types
                summary[error_type] = summary.get(error_type, 0) + 1

            final_output.append({
                "file_name": filename,
                "total_rows": len(details),
                "error_summary": summary,
                "details": details
            })

    return {"reports": final_output}
